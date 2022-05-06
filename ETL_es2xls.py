from elasticsearch import Elasticsearch
from elasticsearch.helpers import bulk, scan
from elasticsearch_dsl import Search
from time import sleep
import openpyxl
from db_restart import Topic_test, File_test
from peewee import *
 
'''
class EsToExcel(object):
    def __init__(self, es_host, index, type, file_name):
        """
        :param es_host: es连接地址 例: 127.0.0.1:9200
        :param index: 索引名
        :param type: type名
        :param file_name: 导出文件名
        """
        self.es = Elasticsearch([es_host])
        self.index = index
        self.type = type
        self.file_name = file_name
        self.item_list = []
 
    def get_field(self):
        """
        获取字段信息
        :return: 字段信息
        """
        field_list = list(self.es.indices.get_mapping(index=self.index, doc_type=self.type)
                          [self.index]['mappings'][self.type]['properties'])
        return field_list
 
    def get_data(self):
        """
        获取索引下所有数据
        :return: 数据纪录
        """
        query = {
            "query": {
                "match_all": {
 
                }
            }
        }
        scanResp = helpers.scan(client=self.es, query=query, scroll="5m", index=self.index, doc_type=self.type,
                                request_timeout=100)
        for resp in scanResp:
            item = resp['_source']
            yield item
 
    def deal_data(self, item, field_list):
        """
        :param item: es查询出的数据纪录
        :param field_list: 对应的字段信息
        :return: 处理好的数据纪录
        """
        new_item = [str(item.get(field, '')) for field in field_list]
        return new_item
 
    def write_excel(self, item_list, style):
        """
        写入Excel
        :param item_list: 数据集
        :param style: 表头信息
        :return:
        """
        f = openpyxl.Workbook()
        sheet1 = f.active
        for index, i in enumerate(style):
            # openpyxl生成xlsx文件，行列从1开始
            sheet1.cell(row=1, column=index + 1, value=i)
        for row, item in enumerate(item_list):
            try:
                for index, i in enumerate(item):
                    # openpyxl生成xlsx文件，行列从1开始
                    sheet1.cell(row=row + 2, column=index + 1, value=i)
            except Exception as e:
                print(e)
 
        f.save('{}.xlsx'.format(self.file_name))
        print('{}数据导出完成'.format(self.file_name))
 
    def start(self):
        """运行主函数"""
        field_list = self.get_field()
        for item in self.get_data():
            new_item = self.deal_data(item, field_list)
            self.item_list.append(new_item)
            print(new_item)
        #self.write_excel(self.item_list, field_list)
 
 
if __name__ == '__main__':
    export = EsToExcel(es_host='127.0.0.1:9200', index='bank', type='text', file_name='test.xlsx')
    export.start()
 '''
 
es = Elasticsearch(hosts="http://127.0.0.1:9200/")

db = SqliteDatabase('zsxq.db')
db.connect()

query_all = {
  "query": {
    "match_all": {}
    }
}

query_match = {
    "query": {
        "match": {
        "text":"新能源"
        }
    },
    'aggs': {
        'by_topic': {'terms': {'field': 'create_time'}}
    },
    'aggs': {
        'doc_count': {'count': {'field': '_source'}}
    }
}

'''
  "sort": [
  {
    "topic_id": {"order": "desc"}
  },
  {
    "para_index": {"order": "asc"}
  }
  ]'''

request_body = {
	"settings": {
       "index": {
            "max_ngram_diff": 2
      },      
        "analysis": {
            "analyzer": {
                "charSplit": {
                    "type": "custom",
                    "tokenizer": "ngram"
                }
            },    
            "tokenizer": { 
                "ngram": {
                    "type": "ngram",
                    "min_gram": 2,
                    "max_gram": 4,
                    "token_chars": [
                        "letter",
                        "digit"
                    ]
                }
            }
        }
    },
    "mappings": { 
      "properties": {
      "text": {
        "type": "text",
        "analyzer": "charSplit",
        "search_analyzer": "charSplit"
        }
      }
    }
}

data = [
    {
        '_index': "bank",
        '_type': "_doc",
        '_source': {"text": value}
    }
    for value in range(5)
]

def handle_es():
    for index, topic in enumerate(data_dic_list):
        yield {
            '_index': "bank",
            '_id': index,
#            '_type': "_doc",
            '_source': {"text": topic['para'], "para_index": topic['para_id'], "topic_id": topic['topic_id'], "topic_time": topic['create_time']}
        }

data_dic_list = []

for row in Topic_test.select().limit(50):
    impact_text = row.topic_content.split("\n")
    impact_text = [x for x in impact_text if x!='']
    for index, each_para in enumerate(impact_text, start=1):
        new_dic = dict(topic_id=str(row.topic_id), para_id=index, para=each_para, create_time = row.create_time)
        data_dic_list.append(new_dic)
db.close()

es.indices.delete(index='bank', ignore=[400, 404])
es.indices.create(index = 'bank', body = request_body)
bulk(es, actions = handle_es(), refresh="wait_for")

#以下为DSL
s = Search(using=es).query("match", text="新能源").sort('-topic_time', 'para_index').extra(size=30).execute()
s_to_dict = s.to_dict()["hits"]["hits"]

bucket_result = Search(using=es).query("match", text="新能源").aggs.bucket('by_topic', 'terms', field='create_time', size=0).execute()
print(bucket_result)
exit()

for each in s:
#    print(each["highlight"]["text"][0])
    print(each["_source"]["topic_time"]+'\t'+str(each["_source"]['para_index']))
exit()

scanResp = scan(client=es, query=query2, scroll="1m", index='bank', request_timeout=100)

title_list = []
f = openpyxl.Workbook()
sheet1 = f.active
for record in scanResp:
    title_list.append(record["_source"]["title"])

# openpyxl生成xlsx文件，行列从1开始
sheet1.cell(row=1, column=1, value='序号')
sheet1.cell(row=1, column=2, value='Title')

for row, item in enumerate(title_list, start=1):
    try:
        # openpyxl生成xlsx文件，行列从1开始
        sheet1.cell(row=row, column=1, value=row)
        sheet1.cell(row=row, column=2, value=item)
    except Exception as e:
        print(e)

f.save('{}.xlsx'.format('test'))